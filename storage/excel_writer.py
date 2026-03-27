# ─────────────────────────────────────────────
#  storage/excel_writer.py
#  Write scraped data to styled Excel files.
# ─────────────────────────────────────────────

import os
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
import config

log = logging.getLogger(__name__)

# ── Palette ───────────────────────────────────────────────────────────────────
CLR_HEADER_RATED   = "1A1A2E"   # deep navy
CLR_HEADER_FONT    = "FFFFFF"
CLR_ROW_ODD        = "F0F4FF"
CLR_ROW_EVEN       = "FFFFFF"
CLR_HEADER_UNRATED = "4A0E0E"   # deep red
CLR_ACCENT         = "E63946"


def _header_style(ws, row: int, headers: list[str], bg: str = CLR_HEADER_RATED):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(bold=True, color=CLR_HEADER_FONT, size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        bottom=Side(style="medium", color="CCCCCC")
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill   = fill
        cell.font   = font
        cell.alignment = align
        cell.border = border
    ws.row_dimensions[row].height = 28


def _row_style(ws, row: int, ncols: int, odd: bool):
    bg = CLR_ROW_ODD if odd else CLR_ROW_EVEN
    fill = PatternFill("solid", fgColor=bg)
    align = Alignment(vertical="center", wrap_text=True)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.alignment = align
    ws.row_dimensions[row].height = 18


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _sheet_title(ws, title: str):
    ws.sheet_properties.tabColor = CLR_ACCENT


# ── General data sheet ────────────────────────────────────────────────────────
GENERAL_HEADERS = [
    "S.No", "Title", "Release Date", "Month/Year",
    "Tomatometer (%)", "Link"
]
GENERAL_WIDTHS = [6, 42, 16, 14, 16, 55]


def write_general(wb: openpyxl.Workbook, data: list[dict], sheet_name: str):
    ws = wb.create_sheet(sheet_name)
    _sheet_title(ws, sheet_name)
    ws.freeze_panes = "A2"

    _header_style(ws, 1, GENERAL_HEADERS,
                  bg=CLR_HEADER_RATED if "Rated" not in sheet_name else CLR_HEADER_RATED)
    _set_col_widths(ws, GENERAL_WIDTHS)

    for i, row in enumerate(data, 1):
        ws.append([
            row.get("s_no", i),
            row.get("title", ""),
            row.get("release_date", ""),
            row.get("release_month", ""),
            row.get("tomatometer", ""),
            row.get("link", ""),
        ])
        _row_style(ws, i + 1, len(GENERAL_HEADERS), odd=(i % 2 == 1))


# ── Insights sheet ────────────────────────────────────────────────────────────
INSIGHTS_HEADERS = [
    "Title", "Synopsis", "Genre", "Director", "Producer",
    "Tomatometer (%)", "Audience Score (%)",
    "Rating", "Runtime", "Release Date (Insights)",
    "Cast & Crew", "Link"
]
INSIGHTS_WIDTHS = [30, 60, 22, 25, 25, 16, 18, 12, 12, 22, 70, 55]


def write_insights(wb: openpyxl.Workbook, data: list[dict], sheet_name: str, bg=CLR_HEADER_RATED):
    ws = wb.create_sheet(sheet_name)
    _sheet_title(ws, sheet_name)
    ws.freeze_panes = "A2"

    _header_style(ws, 1, INSIGHTS_HEADERS, bg=bg)
    _set_col_widths(ws, INSIGHTS_WIDTHS)

    for i, row in enumerate(data, 1):
        ws.append([
            row.get("title", ""),
            row.get("synopsis", ""),
            row.get("genre", ""),
            row.get("director", ""),
            row.get("producer", ""),
            row.get("tomatometer", ""),
            row.get("audience_score", ""),
            row.get("rating", ""),
            row.get("runtime", ""),
            row.get("release_date_(theaters)", row.get("release_date_(streaming)", "")),
            row.get("cast_and_crew", ""),
            row.get("link", ""),
        ])
        _row_style(ws, i + 1, len(INSIGHTS_HEADERS), odd=(i % 2 == 1))


# ── Unrated sheet ─────────────────────────────────────────────────────────────
UNRATED_HEADERS = [
    "S.No", "Title", "Release Date", "Month/Year", "Link", "Reason"
]
UNRATED_WIDTHS = [6, 42, 16, 14, 55, 30]


def write_unrated(wb: openpyxl.Workbook, data: list[dict], sheet_name: str = config.UNRATED_SHEET):
    ws = wb.create_sheet(sheet_name)
    _sheet_title(ws, sheet_name)
    ws.freeze_panes = "A2"

    _header_style(ws, 1, UNRATED_HEADERS, bg=CLR_HEADER_UNRATED)
    _set_col_widths(ws, UNRATED_WIDTHS)

    for i, row in enumerate(data, 1):
        ws.append([
            row.get("s_no", i),
            row.get("title", ""),
            row.get("release_date", ""),
            row.get("release_month", ""),
            row.get("link", ""),
            "No Tomatometer score available",
        ])
        _row_style(ws, i + 1, len(UNRATED_HEADERS), odd=(i % 2 == 1))


# ── Main save function ────────────────────────────────────────────────────────
def save_excel(
    rated_general   : list[dict],
    unrated_general : list[dict],
    rated_insights  : list[dict] | None = None,
    unrated_insights: list[dict] | None = None,
) -> tuple[str, str]:
    """
    Build and save two Excel files:
      - RT_Movies_Rated.xlsx   (rated movies + optional insights)
      - RT_Movies_Unrated.xlsx (unrated movies)
    Returns (rated_path, unrated_path).
    """
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    # ── Rated workbook ────────────────────────────────────────────────────
    wb_rated = openpyxl.Workbook()
    wb_rated.remove(wb_rated.active)  # remove default empty sheet

    write_general(wb_rated, rated_general, config.GENERAL_SHEET)

    if rated_insights:
        write_insights(wb_rated, rated_insights, config.INSIGHTS_SHEET)

    rated_path = os.path.join(config.OUTPUT_DIR, f"RT_Movies_Rated_{ts}.xlsx")
    wb_rated.save(rated_path)
    log.info(f"Saved rated workbook → {rated_path}")

    # ── Unrated workbook ──────────────────────────────────────────────────
    all_unrated = list(unrated_general or [])
    if unrated_insights:
        # merge & deduplicate by title
        seen_titles = {r.get("title","").lower() for r in all_unrated}
        for r in unrated_insights:
            if r.get("title","").lower() not in seen_titles:
                all_unrated.append(r)

    wb_unrated = openpyxl.Workbook()
    wb_unrated.remove(wb_unrated.active)

    if all_unrated:
        write_unrated(wb_unrated, all_unrated)
        if unrated_insights:
            write_insights(wb_unrated, [r for r in all_unrated if r.get("synopsis")],
                           "Unrated Insights", bg=CLR_HEADER_UNRATED)
    else:
        ws = wb_unrated.create_sheet("Not Rated Yet")
        ws["A1"] = "No unrated movies found in this scrape."

    unrated_path = os.path.join(config.OUTPUT_DIR, f"RT_Movies_Unrated_{ts}.xlsx")
    wb_unrated.save(unrated_path)
    log.info(f"Saved unrated workbook → {unrated_path}")

    return rated_path, unrated_path
